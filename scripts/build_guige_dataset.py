#!/usr/bin/env python3
"""Generate data/dataset.json for 规格 (guige) row-level standardization tasks."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
GUIGE_DIR = DATA_DIR / "规格"

ROW_INSTRUCTION_TEMPLATE = """请对以下药品规格原文进行规格标化（单行任务）。

规格(原)：{raw_value}

要求：
1. 按已加载的 guige 技能（规格标化 SOP）将上述原文标化为标准「规格规范」字符串。
2. 将标化结果写入 output.xlsx 的工作表 sheet 的 B2 单元格；A 列原始值保持不变。
3. 若一条原文需拆成多条规格，在 B2 同一单元格内用英文分号「;」连接（不要拆成多行）。
4. 若无法按 SOP 自动判定，在 B2 保留原文并在末尾标注「需人工核查」及简要原因。
5. 不要修改 A2 以外的单元格。"""


def build_row_instruction(raw_value: str) -> str:
    text = str(raw_value).strip() if raw_value is not None else ""
    return ROW_INSTRUCTION_TEMPLATE.format(raw_value=text)


def _load_batch_rows(batch_dir: Path) -> dict:
    init_files = sorted(batch_dir.glob("*_init.xlsx"))
    golden_files = sorted(batch_dir.glob("*_golden.xlsx"))
    if not init_files or not golden_files:
        raise FileNotFoundError(f"Missing *_init.xlsx or *_golden.xlsx in {batch_dir}")

    init_path = init_files[0]
    golden_path = golden_files[0]

    init_wb = openpyxl.load_workbook(init_path, read_only=True, data_only=True)
    golden_wb = openpyxl.load_workbook(golden_path, read_only=True, data_only=True)
    init_ws = init_wb.active
    golden_ws = golden_wb.active

    rows: list[tuple[int, str]] = []
    for row_idx in range(1, init_ws.max_row + 1):
        value = init_ws.cell(row_idx, 1).value
        if value is None or str(value).strip() == "":
            continue
        rows.append((row_idx, str(value).strip()))

    info = {
        "init_file": init_path.name,
        "golden_file": golden_path.name,
        "init_sheet": init_ws.title,
        "golden_sheet": golden_ws.title,
        "row_count": len(rows),
        "rows": rows,
    }
    init_wb.close()
    golden_wb.close()
    return info


def build_dataset(
    *,
    start_row: int = 1,
    end_row: int | None = None,
    batch_filter: str | None = None,
) -> list[dict]:
    if not GUIGE_DIR.is_dir():
        raise FileNotFoundError(f"Guige data directory not found: {GUIGE_DIR}")

    entries: list[dict] = []
    for batch_dir in sorted(GUIGE_DIR.iterdir()):
        if not batch_dir.is_dir():
            continue
        batch_id = batch_dir.name
        if batch_filter and batch_id != batch_filter:
            continue

        info = _load_batch_rows(batch_dir)
        spreadsheet_path = f"规格/{batch_id}"

        for row_idx, raw_value in info["rows"]:
            if row_idx < start_row:
                continue
            if end_row is not None and row_idx > end_row:
                continue

            instance_id = f"{batch_id}-{row_idx:04d}"
            entries.append({
                "id": instance_id,
                "batch_id": batch_id,
                "source_row": row_idx,
                "golden_row": row_idx,
                "instruction": build_row_instruction(raw_value),
                "spreadsheet_path": spreadsheet_path,
                "instruction_type": "Cell-Level Manipulation",
                "answer_position": "B2",
                "answer_sheet": "sheet",
                "golden_sheet": info["golden_sheet"],
                "task_type": "guige_row",
                "init_file": info["init_file"],
                "golden_file": info["golden_file"],
                "raw_value": raw_value,
            })

    return entries


def main() -> None:
    parser = argparse.ArgumentParser(description="Build row-level guige dataset.json")
    parser.add_argument("--output", type=Path, default=DATA_DIR / "dataset.json")
    parser.add_argument("--batch", type=str, default=None, help="Only include this batch id")
    parser.add_argument("--start-row", type=int, default=1)
    parser.add_argument("--end-row", type=int, default=None)
    args = parser.parse_args()

    dataset = build_dataset(
        start_row=args.start_row,
        end_row=args.end_row,
        batch_filter=args.batch,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as fh:
        json.dump(dataset, fh, ensure_ascii=False, indent=2)

    batches = {item["batch_id"] for item in dataset}
    print(f"Wrote {len(dataset)} instance(s) to {args.output}")
    for batch_id in sorted(batches):
        count = sum(1 for item in dataset if item["batch_id"] == batch_id)
        print(f"  - {batch_id}: {count} rows")


if __name__ == "__main__":
    main()
